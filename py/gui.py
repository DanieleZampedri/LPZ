"""
gui.py — Interfaccia grafica per getFV.py
"""

import queue
import sys
import threading
import time
import tkinter as tk
from pathlib import Path
from tkinter import ttk, messagebox
from typing import Optional
import random

HERE = Path(__file__).parent.resolve()
sys.path.insert(0, str(HERE))
import getFV as core
from openpyxl import load_workbook

# ── palette ─────────────────────────────────────────────────────────────────
BG        = "#0f1117"
SURFACE   = "#1a1d27"
SURFACE2  = "#212435"
BORDER    = "#2a2d3a"
GREEN     = "#00c896"
GREEN_DIM = "#00856a"
RED       = "#ff5c5c"
TEXT      = "#e8eaf0"
MUTED     = "#6b7280"
TICKER_FG = "#a5b4fc"

FT = ("Helvetica Neue", 15, "bold")
FC = ("Helvetica Neue", 28, "bold")
FK = ("Helvetica Neue", 12)
FS = ("Helvetica Neue", 10)
FM = ("Menlo", 10)
FB = ("Helvetica Neue", 12, "bold")
FL = ("Helvetica Neue", 11)

# ── lettura fogli disponibili ────────────────────────────────────────────────

def get_available_sheets() -> list[str]:
    """Restituisce i fogli del file Excel escludendo quelli in SHEETS_TO_SKIP."""
    if not core.EXCEL_FILE_PATH.exists():
        return []
    try:
        wb = load_workbook(core.EXCEL_FILE_PATH, read_only=True)
        sheets = [s for s in wb.sheetnames if s not in core.SHEETS_TO_SKIP]
        wb.close()
        return sheets
    except Exception:
        return []

# ── conteggio ticker per i fogli selezionati ─────────────────────────────────

def count_tickers_in_sheets(selected: list[str]) -> tuple[int, str]:
    if not core.EXCEL_FILE_PATH.exists():
        return 0, f"File non trovato:\n{core.EXCEL_FILE_PATH}"
    try:
        wb    = load_workbook(core.EXCEL_FILE_PATH, read_only=True)
        total = 0
        for name in selected:
            if name not in wb.sheetnames:
                continue
            for row in wb[name].iter_rows(min_row=5, min_col=3, max_col=3, values_only=True):
                if row[0] is not None and str(row[0]).strip():
                    total += 1
        wb.close()
        return total, ""
    except Exception as e:
        return 0, str(e)

# ── worker ───────────────────────────────────────────────────────────────────

def run_worker(
    q: queue.Queue,
    stop_event: threading.Event,
    selected_sheets: list[str],
    dry_run: bool = False,
) -> None:
    def put(*args):
        q.put(args)

    if not core.COOKIE or not core.COOKIE.strip():
        put("error", "Cookie non trovato.\nImposta COOKIE nel file .env")
        return

    if not core.EXCEL_FILE_PATH.exists():
        put("error", f"File Excel non trovato:\n{core.EXCEL_FILE_PATH}")
        return

    try:
        session = core.create_session()
        wb      = load_workbook(core.EXCEL_FILE_PATH)

        for sheet_name in wb.sheetnames:
            if stop_event.is_set():
                break
            # salta se non selezionato o in SHEETS_TO_SKIP
            if sheet_name not in selected_sheets:
                continue

            ws     = wb[sheet_name]
            offset = core.SHEET_COLUMN_OFFSET.get(sheet_name, 0)

            tickers: list[tuple[int, str]] = []
            for idx, row in enumerate(
                ws.iter_rows(min_row=5, min_col=3, max_col=3, values_only=True), start=5
            ):
                if row[0] is not None and str(row[0]).strip():
                    tickers.append((idx, str(row[0]).strip()))

            if not tickers:
                continue

            put("status", f"Sheet: {sheet_name}  —  {len(tickers)} ticker")

            for b_start in range(0, len(tickers), core.BATCH_SIZE):
                if stop_event.is_set():
                    break
                for row_num, ticker in tickers[b_start: b_start + core.BATCH_SIZE]:
                    if stop_event.is_set():
                        break
                    put("ticker_start", ticker)
                    try:
                        result = core.extract_all(ticker, session)
                    except Exception as e:
                        result = None
                        put("log_err", f"[{ticker}] errore imprevisto: {e}")

                    if result is not None:
                        core.write_result(ws, row_num, result, offset, dry_run)
                        put("ticker_done", ticker, True)
                    else:
                        put("ticker_done", ticker, False)

                    time.sleep(random.uniform(core.TICKER_DELAY_MIN, core.TICKER_DELAY_MAX))

                if not dry_run and not stop_event.is_set():
                    wb.save(core.EXCEL_FILE_PATH)

                if b_start + core.BATCH_SIZE < len(tickers) and not stop_event.is_set():
                    wait = random.uniform(core.BATCH_WAIT_MIN, core.BATCH_WAIT_MAX)
                    put("status", f"Pausa batch — {wait:.0f}s...")
                    time.sleep(wait)

            if not dry_run and not stop_event.is_set():
                wb.save(core.EXCEL_FILE_PATH)

        put("done", None)

    except Exception as e:
        put("error", str(e))

# ── GUI ──────────────────────────────────────────────────────────────────────

class App(tk.Tk):

    def __init__(self):
        super().__init__()
        self.title("LPZ Investing")
        self.configure(bg=BG)
        self.resizable(False, False)

        self._q      = queue.Queue()
        self._stop   = threading.Event()
        self._thread: Optional[threading.Thread] = None

        self._total  = 0
        self._done   = 0
        self._errors = 0
        self._t0: Optional[float] = None

        # variabili checkbox fogli
        self._sheet_vars: dict[str, tk.BooleanVar] = {}
        self._select_all_var = tk.BooleanVar(value=True)

        self._build()
        self._load_sheets()
        self._poll()

    # ── costruzione UI ───────────────────────────────────────────────────────

    def _build(self):
        self.geometry("520x760")

        f = tk.Frame(self, bg=BG)
        f.pack(fill="both", expand=True, padx=28, pady=20)

        # header
        hdr = tk.Frame(f, bg=BG)
        hdr.pack(fill="x", pady=(0, 16))
        tk.Label(hdr, text="LPZ Investing", font=FT, bg=BG, fg=TEXT).pack(side="left")
        self._badge = tk.Label(hdr, text="pronto", font=FS,
                               bg=SURFACE, fg=MUTED, padx=10, pady=4)
        self._badge.pack(side="right")

        # ── sezione selezione fogli ──
        sheets_card = tk.Frame(f, bg=SURFACE, highlightbackground=BORDER,
                               highlightthickness=1)
        sheets_card.pack(fill="x", pady=(0, 12))

        sheets_header = tk.Frame(sheets_card, bg=SURFACE)
        sheets_header.pack(fill="x", padx=16, pady=(12, 6))

        tk.Label(sheets_header, text="Fogli da aggiornare",
                 font=("Helvetica Neue", 11, "bold"),
                 bg=SURFACE, fg=TEXT).pack(side="left")

        # checkbox seleziona tutto
        cb_all = tk.Checkbutton(
            sheets_header, text="Seleziona tutto",
            variable=self._select_all_var,
            font=FS, bg=SURFACE, fg=MUTED,
            activebackground=SURFACE, activeforeground=TEXT,
            selectcolor=SURFACE2,
            relief="flat", cursor="hand2",
            command=self._on_select_all
        )
        cb_all.pack(side="right")

        # separatore
        tk.Frame(sheets_card, bg=BORDER, height=1).pack(fill="x", padx=16)

        # contenitore scrollabile per i fogli
        self._sheets_frame = tk.Frame(sheets_card, bg=SURFACE)
        self._sheets_frame.pack(fill="x", padx=16, pady=(6, 12))

        # ── card contatore ──
        card = tk.Frame(f, bg=SURFACE, highlightbackground=BORDER,
                        highlightthickness=1)
        card.pack(fill="x", pady=(0, 12))
        inner = tk.Frame(card, bg=SURFACE)
        inner.pack(padx=20, pady=14)
        self._lbl_count = tk.Label(inner, text="— / —", font=FC,
                                   bg=SURFACE, fg=TEXT)
        self._lbl_count.pack()
        self._lbl_ticker = tk.Label(inner, text="In attesa di avvio…",
                                    font=FK, bg=SURFACE, fg=TICKER_FG)
        self._lbl_ticker.pack(pady=(4, 0))

        # progressbar
        bf = tk.Frame(f, bg=BG)
        bf.pack(fill="x", pady=(0, 6))
        sty = ttk.Style(self)
        sty.theme_use("clam")
        sty.configure("lpz.Horizontal.TProgressbar",
                       troughcolor=SURFACE, background=GREEN,
                       bordercolor=BG, lightcolor=GREEN, darkcolor=GREEN,
                       thickness=6)
        self._bar = ttk.Progressbar(bf, orient="horizontal", length=464,
                                    mode="determinate",
                                    style="lpz.Horizontal.TProgressbar")
        self._bar.pack()

        # tempi
        tf = tk.Frame(f, bg=BG)
        tf.pack(fill="x", pady=(6, 2))
        self._lbl_el  = tk.Label(tf, text="Trascorso  —", font=FS, bg=BG, fg=MUTED)
        self._lbl_rem = tk.Label(tf, text="Rimanente  —", font=FS, bg=BG, fg=MUTED)
        self._lbl_el.pack(side="left")
        self._lbl_rem.pack(side="right")

        # status
        self._lbl_st = tk.Label(f, text="", font=FS, bg=BG, fg=MUTED, anchor="w")
        self._lbl_st.pack(fill="x", pady=(6, 0))

        # log
        lf = tk.Frame(f, bg=SURFACE, highlightbackground=BORDER, highlightthickness=1)
        lf.pack(fill="x", pady=(10, 0))
        self._log = tk.Text(lf, width=56, height=7, font=FM,
                            bg=SURFACE, fg="#9ca3af",
                            relief="flat", state="disabled",
                            wrap="none", padx=14, pady=8)
        self._log.pack(side="left", fill="both", expand=True)
        sb = tk.Scrollbar(lf, orient="vertical", command=self._log.yview,
                          bg=SURFACE, troughcolor=SURFACE, width=8)
        sb.pack(side="right", fill="y")
        self._log.config(yscrollcommand=sb.set)
        self._log.tag_config("ok",  foreground=GREEN)
        self._log.tag_config("err", foreground=RED)
        self._log.tag_config("inf", foreground=MUTED)
        self._log.tag_config("ts",  foreground="#374151")

        # bottoni
        br = tk.Frame(f, bg=BG)
        br.pack(fill="x", pady=(16, 0))
        self._btn_go = tk.Button(
            br, text="Avvia aggiornamento", font=FB,
            bg=GREEN, fg="#0a0f0d",
            activebackground=GREEN_DIM, activeforeground="#0a0f0d",
            relief="flat", padx=20, pady=10, cursor="hand2",
            command=self._on_start
        )
        self._btn_go.pack(side="left", fill="x", expand=True, padx=(0, 8))
        self._btn_stop = tk.Button(
            br, text="Interrompi", font=FB,
            bg=BORDER, fg=MUTED,
            activebackground="#3a3d4a", activeforeground=TEXT,
            relief="flat", padx=20, pady=10, cursor="hand2",
            state="disabled", command=self._on_stop
        )
        self._btn_stop.pack(side="left", fill="x", expand=True)

    # ── caricamento fogli ────────────────────────────────────────────────────

    def _load_sheets(self):
        self._log_ln("Lettura file Excel…", "inf")
        def _work():
            sheets = get_available_sheets()
            self._q.put(("sheets_loaded", sheets))
        threading.Thread(target=_work, daemon=True).start()

    def _populate_sheets(self, sheets: list[str]):
        """Popola il frame con una checkbox per ogni foglio."""
        for widget in self._sheets_frame.winfo_children():
            widget.destroy()
        self._sheet_vars.clear()

        if not sheets:
            tk.Label(self._sheets_frame, text="Nessun foglio trovato",
                     font=FS, bg=SURFACE, fg=RED).pack(anchor="w")
            return

        # griglia a 2 colonne
        for i, sheet in enumerate(sheets):
            var = tk.BooleanVar(value=True)
            self._sheet_vars[sheet] = var
            cb = tk.Checkbutton(
                self._sheets_frame,
                text=sheet,
                variable=var,
                font=FL, bg=SURFACE, fg=TEXT,
                activebackground=SURFACE, activeforeground=TEXT,
                selectcolor=SURFACE2,
                relief="flat", cursor="hand2",
                command=self._on_sheet_toggle
            )
            col = i % 2
            row = i // 2
            cb.grid(row=row, column=col, sticky="w", padx=(0, 20), pady=2)

        # aggiorna contatore
        self._refresh_count_async()

    def _on_select_all(self):
        val = self._select_all_var.get()
        for var in self._sheet_vars.values():
            var.set(val)
        self._refresh_count_async()

    def _on_sheet_toggle(self):
        """Aggiorna lo stato di 'seleziona tutto' in base alle checkbox."""
        all_checked = all(v.get() for v in self._sheet_vars.values())
        none_checked = not any(v.get() for v in self._sheet_vars.values())
        if all_checked:
            self._select_all_var.set(True)
        elif none_checked:
            self._select_all_var.set(False)
        self._refresh_count_async()

    def _refresh_count_async(self):
        """Aggiorna il contatore ticker in background al cambio selezione."""
        selected = self._selected_sheets()
        def _work():
            total, err = count_tickers_in_sheets(selected)
            self._q.put(("count_update", total, err))
        threading.Thread(target=_work, daemon=True).start()

    def _selected_sheets(self) -> list[str]:
        return [s for s, v in self._sheet_vars.items() if v.get()]

    # ── logica avvio/stop ────────────────────────────────────────────────────

    def _on_start(self):
        selected = self._selected_sheets()
        if not selected:
            messagebox.showwarning("Attenzione", "Seleziona almeno un foglio.")
            return
        if self._total == 0:
            messagebox.showwarning(
                "Attenzione",
                f"Nessun ticker trovato nei fogli selezionati.\n"
                f"Percorso: {core.EXCEL_FILE_PATH}"
            )
            return

        self._stop.clear()
        self._done = self._errors = 0
        self._t0 = time.time()
        self._bar["value"]   = 0
        self._bar["maximum"] = self._total
        self._badge.config(text="in corso", fg=GREEN)
        self._btn_go.config(state="disabled")
        self._btn_stop.config(state="normal", bg="#3a1a1a", fg=RED)
        self._log_ln(f"Avvio su {len(selected)} fogli…", "inf")

        self._thread = threading.Thread(
            target=run_worker,
            args=(self._q, self._stop, selected),
            daemon=True
        )
        self._thread.start()

    def _on_stop(self):
        self._stop.set()
        self._btn_stop.config(state="disabled")
        self._log_ln("Interruzione in corso…", "inf")

    # ── poll & dispatch ──────────────────────────────────────────────────────

    def _poll(self):
        try:
            while True:
                self._dispatch(self._q.get_nowait())
        except queue.Empty:
            pass
        self.after(80, self._poll)

    def _dispatch(self, msg: tuple):
        k = msg[0]

        if k == "sheets_loaded":
            sheets = msg[1]
            self._populate_sheets(sheets)
            self._log_ln(f"{len(sheets)} fogli disponibili.", "inf")

        elif k == "count_update":
            _, total, err = msg
            self._total = total
            if err:
                self._lbl_count.config(text="Errore")
                self._log_ln(f"Errore conteggio: {err}", "err")
            else:
                self._lbl_count.config(text=f"0 / {total}")

        elif k == "ticker_start":
            self._lbl_ticker.config(text=msg[1])

        elif k == "ticker_done":
            _, ticker, ok = msg
            self._done += 1
            if not ok:
                self._errors += 1
            self._bar["value"] = self._done
            self._lbl_count.config(text=f"{self._done} / {self._total}")
            self._log_ln(f"{'✓' if ok else '✗'}  {ticker}", "ok" if ok else "err")
            self._tick_time()

        elif k == "status":
            self._lbl_st.config(text=msg[1])

        elif k == "log_err":
            self._log_ln(msg[1], "err")

        elif k == "done":
            self._finish(False)

        elif k == "error":
            self._log_ln(f"ERRORE: {msg[1]}", "err")
            messagebox.showerror("Errore", msg[1])
            self._finish(True)

    def _finish(self, interrupted: bool):
        self._btn_go.config(state="normal")
        self._btn_stop.config(state="disabled", bg=BORDER, fg=MUTED)
        self._badge.config(
            text="interrotto" if interrupted else "completato", fg=MUTED
        )
        self._lbl_ticker.config(text="Operazione terminata")
        summary = f"{self._done} elaborati  ·  {self._errors} errori"
        self._lbl_st.config(text=summary)
        self._log_ln(f"— {summary} —", "inf")

    def _tick_time(self):
        if not self._t0 or self._done == 0:
            return
        el  = time.time() - self._t0
        rem = (el / self._done) * (self._total - self._done)
        self._lbl_el.config( text=f"Trascorso   {self._fmt(el)}")
        self._lbl_rem.config(text=f"Rimanente   {self._fmt(rem)}")

    @staticmethod
    def _fmt(s: float) -> str:
        h, r = divmod(int(s), 3600)
        m, s = divmod(r, 60)
        if h: return f"{h}h {m:02d}m"
        if m: return f"{m}m {s:02d}s"
        return f"{s}s"

    def _log_ln(self, text: str, tag: str = ""):
        self._log.config(state="normal")
        ts = time.strftime("%H:%M:%S")
        self._log.insert("end", f"[{ts}]  ", "ts")
        self._log.insert("end", f"{text}\n", tag)
        self._log.see("end")
        self._log.config(state="disabled")


# ── entrypoint ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    App().mainloop()