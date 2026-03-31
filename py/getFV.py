import argparse
import logging
import random
import sys
import time
from pathlib import Path
from typing import Optional

import requests
from fxconverter import CurrencyConverter
from openpyxl import load_workbook
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from dotenv import load_dotenv
import os


EXCEL_FILE_PATH = Path("../excel/Valuation Model LPZ Investing.xlsx")
LOG_FILE = Path("getFV.log")

load_dotenv()
COOKIE: str = os.getenv("COOKIE", "")

URL = "https://it.investing.com/pro/_/api/query?raw=true"

HEADERS = {
    "Content-Type": "application/json",
    "User-Agent": "Mozilla/5.0",
    "Referer": "https://it.investing.com/",
    "Cookie": COOKIE,
}

GRAPHQL_QUERY = """
query loadModel ($slug: String!, $ticker: String!) {
  model: create_model (slug: $slug, ticker: $ticker) {
    workbook
  }
}
"""

# Slug dei due workbook
SLUG_DCF    = "dcf-growth-exit-5yr"
SLUG_EBITDA = "ebitda-multiples"

# Fogli da saltare
SHEETS_TO_SKIP: set[str] = {
    "TOTAL SCORES",
    "Set up",
    "Company names",
    "Database"
}

# Offset di colonna per-sheet (0 = nessun offset)
SHEET_COLUMN_OFFSET: dict[str, int] = {
    "Real Estate": 1,
}

# Colonne di destinazione nel foglio Excel (1-based, senza offset)
# ev_ebitda è in colonna 9, subito dopo fv (colonna 8)
COLUMN_MAP = {
    "fv":         8,
    "ev_ebitda":  12,
    "ebitda_gp":  26,
    "ebitda_gnp": 29,
    "fcf":        32,
    "net_debt":   35,
}


# Celle del workbook DCF da leggere
DCF_SHEET = "5 Year DCF - Growth Exit"
CELLS = {
    "ebitda_gpn_v1": "E368",
    "ebitda_gpn_v5": "I368",
}

EBITDA_GP_V1_CELLS = [f"D{r}" for r in range(256, 269)] + ["D254","D96"]
EBITDA_GP_V5_CELLS = [f"I{r}" for r in range(256, 269)] + ["I254", "I96"]
NET_DEBT_DEBT_CELL = "E224"
NET_DEBT_CASH_CELL = "E222"

BATCH_SIZE            = 30
MAX_ATTEMPTS          = 3
RATE_LIMIT_BASE_WAIT  = 30   # secondi base in caso di 429

BATCH_WAIT_MIN        = 4   # attesa tra un batch e il successivo
BATCH_WAIT_MAX        = 6

TICKER_DELAY_MIN      = 0.5 # attesa tra un ticker e il successivo
TICKER_DELAY_MAX      = 0.7

INTRA_TICKER_DELAY_MIN = 0.3  # attesa tra chiamata DCF e chiamata EV/EBITDA
INTRA_TICKER_DELAY_MAX = 0.6  # dello stesso ticker

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------

def setup_logging() -> logging.Logger:
    logger = logging.getLogger("getFV")
    logger.setLevel(logging.DEBUG)

    fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", datefmt="%Y-%m-%d %H:%M:%S")

    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(fmt)

    fh = logging.FileHandler(LOG_FILE, encoding="utf-8", mode="w")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(fmt)

    logger.addHandler(ch)
    logger.addHandler(fh)
    return logger


logger = setup_logging()

def validate_cookie() -> None:
    """Verifica che il cookie sia presente; termina lo script altrimenti."""
    if not COOKIE or COOKIE.strip() == "":
        logger.error(
            "Cookie non trovato. Imposta la variabile d'ambiente COOKIE nel file .env."
        )
        sys.exit(1)
    logger.debug("Cookie trovato (lunghezza %d caratteri).", len(COOKIE))


def create_session() -> requests.Session:
    session = requests.Session()
    retry_strategy = Retry(
        total=2,
        backoff_factor=1,
        status_forcelist=[500, 502, 503, 504],
        allowed_methods=["POST"],
    )
    adapter = HTTPAdapter(max_retries=retry_strategy)
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    return session


_converter = CurrencyConverter()

def to_usd(amount: float, currency: str) -> float:
    if currency.upper() == "USD":
        return float(amount)
    try:
        return _converter.convert(amount, currency.upper(), "USD")
    except Exception as e:
        logger.warning("Impossibile convertire %s -> USD: %s. Valore non convertito.", currency, e)
        return float(amount)


def cagr(v_start: Optional[float], v_end: Optional[float], years: int = 4) -> Optional[float]:
    if v_start is None or v_end is None or isinstance(v_start, str) or isinstance(v_end, str):
        return None
    if v_start == 0:
        return None
    rate = ((v_end / v_start) ** (1 / years) - 1)
    if isinstance(rate, complex) or isinstance(rate, str):
        return None
    try:
        return rate
    except (ZeroDivisionError, ValueError, OverflowError) as e:
        logger.debug("CAGR non calcolabile: %s (v_start=%s, v_end=%s).", e, v_start, v_end)
        return None

def sum_cells(dcf_cells: dict, cell_ids: list[str]) -> Optional[float]:
    total     = 0.0
    found_any = False
    for cell_id in cell_ids:
        if cell_id in dcf_cells:
            val = dcf_cells[cell_id].get("value")
            if val is not None and not isinstance(val, str):
                try:
                    total += float(val)
                    found_any = True
                except (TypeError, ValueError):
                    logger.debug("Cella %s: valore non numerico (%s), ignorato.", cell_id, val)
    return total if found_any else None

def calc_net_debt(dcf_cells: dict, ebitda_gp_v5: Optional[float]) -> Optional[float]:
    if ebitda_gp_v5 is None or ebitda_gp_v5 == 0:
        return None
    debt = dcf_cells.get(NET_DEBT_DEBT_CELL, {}).get("value")
    cash = dcf_cells.get(NET_DEBT_CASH_CELL, {}).get("value")
    if debt is None or cash is None:
        logger.debug("calc_net_debt: cella mancante (debt=%s, cash=%s)", 
                     NET_DEBT_DEBT_CELL, NET_DEBT_CASH_CELL)
        return None
    if isinstance(debt, str) or isinstance(cash, str):
        return None
    try:
        logger.debug("Calcolo net debt: debt=%s, cash=%s, ebitda_gp_v5=%s", -debt, cash, ebitda_gp_v5)
        return (-float(debt) - float(cash)) / ebitda_gp_v5
    except (TypeError, ValueError, ZeroDivisionError) as e:
        logger.debug("calc_net_debt non calcolabile: %s", e)
        return None

def fetch_workbook(ticker: str, slug: str, session: requests.Session) -> Optional[dict]:
   
    payload = {
        "query": GRAPHQL_QUERY,
        "variables": {"slug": slug, "ticker": ticker},
    }

    for attempt in range(MAX_ATTEMPTS):
        try:
            response = session.post(URL, json=payload, headers=HEADERS, timeout=30)

            if response.status_code == 429:
                wait = (2 ** attempt) * RATE_LIMIT_BASE_WAIT
                logger.warning(
                    "[%s/%s] Rate limit. Attesa %d s (tentativo %d/%d)...",
                    ticker, slug, wait, attempt + 1, MAX_ATTEMPTS,
                )
                time.sleep(wait)
                continue

            if response.status_code != 200:
                logger.error("[%s/%s] Errore HTTP %d.", ticker, slug, response.status_code)
                return None

            data = response.json()
            model = (data.get("data") or {}).get("model")
            if model is None:
                logger.error("[%s/%s] Risposta API senza 'model': %s", ticker, slug, data)
                return None

            return model["workbook"]

        except requests.exceptions.RequestException as e:
            logger.error("[%s/%s] Errore di rete (tentativo %d/%d): %s",
                         ticker, slug, attempt + 1, MAX_ATTEMPTS, e)
            if attempt < MAX_ATTEMPTS - 1:
                time.sleep(2 ** attempt * 5)

    return None


def extract_dcf(ticker: str, session: requests.Session) -> Optional[dict]:
    workbook = fetch_workbook(ticker, SLUG_DCF, session)
    if workbook is None:
        return None

    try:
        currency: str = workbook["properties"]["trading_currency"]
        dcf_cells: dict = workbook["sheets"][DCF_SHEET]["cells"]
        named: dict = workbook["named_values"]

        def cell(key: str) -> Optional[float]:
            cell_id = CELLS[key]
            return dcf_cells[cell_id]["value"] if cell_id in dcf_cells else None

        ebitda_gp_v1 = sum_cells(dcf_cells, EBITDA_GP_V1_CELLS)
        ebitda_gp_v5 = sum_cells(dcf_cells, EBITDA_GP_V5_CELLS)

        result = {
            "fv":         named.get("fv_mid"),
            "ebitda_gp":  cagr(ebitda_gp_v1, ebitda_gp_v5),
            "ebitda_gnp": cagr(cell("ebitda_gpn_v1"), cell("ebitda_gpn_v5")),
            "fcf":        named.get("_unlevered_fcf_5y_cagr"),
            "net_debt":   calc_net_debt(dcf_cells, ebitda_gp_v5),
        }

        if currency.upper() != "USD" and result["fv"] is not None:
            result["fv"] = to_usd(result["fv"], currency)

        logger.debug("[%s] DCF: %s (valuta: %s)", ticker, result, currency)
        logger.info(
            "[%s] EBITDA GP  v1=%s  v5=%s  CAGR=%s",
            ticker, ebitda_gp_v1, ebitda_gp_v5,
            cagr(ebitda_gp_v1, ebitda_gp_v5)
        )
        return result

    except KeyError as e:
        logger.error("[%s] KeyError nel workbook DCF: %s", ticker, e)
        return None


def extract_ev_ebitda(ticker: str, session: requests.Session) -> Optional[float]:
    workbook = fetch_workbook(ticker, SLUG_EBITDA, session)
    if workbook is None:
        return None

    try:
        currency: str = workbook["properties"]["trading_currency"]
        named: dict = workbook["named_values"]
        value: Optional[float] = named.get("fv_mid")

        if value is not None and currency.upper() != "USD":
            value = to_usd(value, currency)

        logger.debug("[%s] EV/EBITDA fv_mid: %s (valuta: %s)", ticker, value, currency)
        return value

    except KeyError as e:
        logger.error("[%s] KeyError nel workbook EV/EBITDA: %s", ticker, e)
        return None


def extract_all(ticker: str, session: requests.Session) -> Optional[dict]:
    """
    Esegue in sequenza le due chiamate API per il ticker:
      1. dcf-growth-exit-5yr
      2. [delay casuale INTRA_TICKER]
      3. dcf-ebitda-exit-5yr

    Restituisce un dict con tutti i campi, o None se la chiamata principale fallisce.
    Il campo ev_ebitda puo' essere None se la seconda chiamata fallisce (non blocca).
    """
    # Prima chiamata — DCF principale
    dcf = extract_dcf(ticker, session)
    if dcf is None:
        return None

    # Delay tra le due chiamate dello stesso ticker
    intra_delay = random.uniform(INTRA_TICKER_DELAY_MIN, INTRA_TICKER_DELAY_MAX)
    logger.debug("[%s] Attesa intra-ticker %.1f s prima di EV/EBITDA...", ticker, intra_delay)
    time.sleep(intra_delay)

    # Seconda chiamata — EV/EBITDA
    ev_ebitda = extract_ev_ebitda(ticker, session)
    if ev_ebitda is None:
        logger.warning("[%s] EV/EBITDA non disponibile, verra' lasciato vuoto.", ticker)

    return {**dcf, "ev_ebitda": ev_ebitda}


def write_result(ws, row: int, result: dict, offset: int, dry_run: bool) -> None:
    for field, base_col in COLUMN_MAP.items():
        value = result.get(field)
        if value is not None:
            col = base_col + offset
            if not dry_run:
                ws.cell(row=row, column=col, value=value)
            logger.debug("  [%s] %s -> col %d = %s%s",
                         ws.title, field, col, value,
                         " (dry-run, non scritto)" if dry_run else "")



def process_sheets(dry_run: bool = False) -> None:
    validate_cookie()

    if not EXCEL_FILE_PATH.exists():
        raise FileNotFoundError(f"File Excel non trovato: {EXCEL_FILE_PATH}")

    session = create_session()
    wb = load_workbook(EXCEL_FILE_PATH)

    for sheet_name in wb.sheetnames:
        if sheet_name in SHEETS_TO_SKIP:
            logger.info("Sheet ignorato: %s", sheet_name)
            continue

        logger.info("=== Elaborazione sheet: %s ===", sheet_name)
        ws = wb[sheet_name]
        offset = SHEET_COLUMN_OFFSET.get(sheet_name, 0)

        # Leggi i ticker dalla colonna C (colonna 3), riga 5 in poi (header in riga 4)
        tickers: list[tuple[int, str]] = []
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=5, min_col=3, max_col=3, values_only=True), start=5
        ):
            value = row[0]
            if value is not None and str(value).strip():
                tickers.append((row_idx, str(value).strip()))

        if not tickers:
            logger.warning("Nessun ticker trovato in %s.", sheet_name)
            continue

        logger.info("%d ticker trovati in %s.", len(tickers), sheet_name)

        for batch_start in range(0, len(tickers), BATCH_SIZE):
            batch = tickers[batch_start: batch_start + BATCH_SIZE]

            for row_number, ticker in batch:
                logger.info("  [%s] Elaborazione (riga %d)...", ticker, row_number)
                result = extract_all(ticker, session)

                if result is not None:
                    write_result(ws, row_number, result, offset, dry_run)
                    logger.info("  [%s] Completato.", ticker)
                else:
                    logger.warning("  [%s] Impossibile estrarre dati.", ticker)

                # Delay tra un ticker e il successivo
                time.sleep(random.uniform(TICKER_DELAY_MIN, TICKER_DELAY_MAX))

            # Salva il workbook al termine di ogni batch
            if not dry_run:
                wb.save(EXCEL_FILE_PATH)
                logger.info("Workbook salvato dopo batch (%s).", sheet_name)

            # Attesa tra batch (se non e' l'ultimo)
            if batch_start + BATCH_SIZE < len(tickers):
                wait = random.uniform(BATCH_WAIT_MIN, BATCH_WAIT_MAX)
                logger.info("Attesa %.1f s prima del prossimo batch...", wait)
                time.sleep(wait)

        # Salvataggio finale del sheet
        if not dry_run:
            wb.save(EXCEL_FILE_PATH)
            logger.info("Sheet '%s' completato e salvato.", sheet_name)

    logger.info("Elaborazione completata.")



def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Estrae dati di valutazione da Investing.com Pro e li scrive su Excel."
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Esegue l'estrazione senza scrivere nulla su Excel (utile per test).",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    if args.dry_run:
        logger.info("Modalita' DRY-RUN attiva: nessuna modifica verra' scritta su Excel.")

    try:
        process_sheets(dry_run=args.dry_run)
    except FileNotFoundError as e:
        logger.error("%s", e)
        sys.exit(1)
    except KeyboardInterrupt:
        logger.warning("Esecuzione interrotta dall'utente.")
        sys.exit(0)
    except Exception as e:
        logger.exception("Errore imprevisto: %s", e)
        sys.exit(1)


if __name__ == "__main__":
    main()